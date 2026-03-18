import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("scenarios/risk_mitigation_regression_matrix.json") as f:
    data = json.load(f)

wb = Workbook()

# --- Sheet 1: Scenarios ---
ws1 = wb.active
ws1.title = "Scenarios"

cols = [
    "use_case_id", "poc_demo", "scenario_title", "scenario_type",
    "requirement_ids", "preconditions", "input_payload", "transport",
    "auth_mode", "expected_http_status", "expected_result",
    "expected_events", "error_condition", "test_tags",
]
col_widths = [14, 28, 65, 12, 28, 32, 55, 12, 18, 12, 55, 30, 28, 35]

hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor="4472C4")
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_font = Font(name="Arial", size=9)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

for ci, col in enumerate(cols, 1):
    c = ws1.cell(row=1, column=ci, value=col)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = hdr_align
    c.border = thin_border
    ws1.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]

pos_fill = PatternFill("solid", fgColor="E2EFDA")
neg_fill = PatternFill("solid", fgColor="FCE4EC")
edge_fill = PatternFill("solid", fgColor="FFF3E0")

wrap_cols = {3, 5, 6, 7, 11, 14}
for ri, scenario in enumerate(data, 2):
    for ci, col in enumerate(cols, 1):
        val = scenario.get(col, "")
        if isinstance(val, (list, dict)):
            val = json.dumps(val, ensure_ascii=False)
        c = ws1.cell(row=ri, column=ci, value=val)
        c.font = cell_font
        c.border = thin_border
        c.alignment = Alignment(vertical="top", wrap_text=(ci in wrap_cols))
        st = scenario.get("scenario_type", "")
        if st == "positive":
            c.fill = pos_fill
        elif st == "negative":
            c.fill = neg_fill
        elif st == "edge":
            c.fill = edge_fill

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:N{len(data)+1}"

# --- Sheet 2: Coverage Summary ---
ws2 = wb.create_sheet("Coverage Summary")

tag_type_counts = {}
for s in data:
    st = s["scenario_type"]
    for tag in s["test_tags"]:
        if tag.startswith("mitigation-") or tag.startswith("capability-"):
            if tag not in tag_type_counts:
                tag_type_counts[tag] = {"positive": 0, "negative": 0, "edge": 0, "total": 0}
            tag_type_counts[tag][st] += 1
            tag_type_counts[tag]["total"] += 1

sorted_tags = sorted(
    tag_type_counts.keys(),
    key=lambda x: (0 if x.startswith("mitigation") else 1, x),
)

cov_headers = ["Coverage Area", "Positive", "Negative", "Edge", "Total"]
cov_widths = [30, 12, 12, 12, 12]
for ci, h in enumerate(cov_headers, 1):
    c = ws2.cell(row=1, column=ci, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = hdr_align
    c.border = thin_border
    ws2.column_dimensions[get_column_letter(ci)].width = cov_widths[ci - 1]

for ri, tag in enumerate(sorted_tags, 2):
    counts = tag_type_counts[tag]
    ws2.cell(row=ri, column=1, value=tag).font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=ri, column=2, value=counts["positive"]).font = cell_font
    ws2.cell(row=ri, column=3, value=counts["negative"]).font = cell_font
    ws2.cell(row=ri, column=4, value=counts["edge"]).font = cell_font
    ws2.cell(row=ri, column=5, value=counts["total"]).font = Font(
        name="Arial", bold=True, size=10
    )
    for ci in range(1, 6):
        ws2.cell(row=ri, column=ci).border = thin_border
        ws2.cell(row=ri, column=ci).alignment = Alignment(
            horizontal="center" if ci > 1 else "left"
        )

tr = len(sorted_tags) + 2
ws2.cell(row=tr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=11)
for ci in range(2, 6):
    formula = f"=SUM({get_column_letter(ci)}2:{get_column_letter(ci)}{tr-1})"
    ws2.cell(row=tr, column=ci, value=formula)
    ws2.cell(row=tr, column=ci).font = Font(name="Arial", bold=True, size=11)
    ws2.cell(row=tr, column=ci).border = thin_border

dr = tr + 2
ws2.cell(row=dr, column=1, value="Scenario Distribution").font = Font(
    name="Arial", bold=True, size=11, color="4472C4"
)
types_count = {"positive": 0, "negative": 0, "edge": 0}
for s in data:
    types_count[s["scenario_type"]] += 1
for i, (label, count) in enumerate(types_count.items()):
    pct = round(100 * count / len(data), 1)
    ws2.cell(row=dr + 1 + i, column=1, value=label.capitalize()).font = cell_font
    ws2.cell(row=dr + 1 + i, column=2, value=count).font = cell_font
    ws2.cell(row=dr + 1 + i, column=3, value=f"{pct}%").font = cell_font

ws2.freeze_panes = "A2"

# --- Sheet 3: Regression Case ---
ws3 = wb.create_sheet("Regression Case")
ws3.column_dimensions["A"].width = 25
ws3.column_dimensions["B"].width = 100

title_font = Font(name="Arial", bold=True, size=14, color="4472C4")
section_font = Font(name="Arial", bold=True, size=11, color="4472C4")
body_font = Font(name="Arial", size=10)
section_fill = PatternFill("solid", fgColor="D6E4F0")

rows = [
    ("", "FULL REGRESSION TEST CASE FOR RISK MITIGATIONS AND CAPABILITIES", title_font, None),
    ("", "", body_font, None),
    ("Purpose", "Validate that all 18 risk mitigations and 6 platform capabilities function correctly across the multi-sovereign healthcare AI federation. This regression suite provides a repeatable, auditable evidence base for regulatory compliance (GDPR, HIPAA, NHS DSPT) and platform operational readiness.", section_font, section_fill),
    ("", "", body_font, None),
    ("Scope", "18 Risk Mitigations:", section_font, section_fill),
    ("", "MIT-1.1 Transfer Impact Assessment  |  MIT-1.2 Rate-Limit Circuit Breaker  |  MIT-1.3 Runtime Enforcement Proxy", body_font, None),
    ("", "MIT-2.1 Rule Governance Versioning  |  MIT-2.2 Automated PHI Remediation  |  MIT-2.3 Multi-Layer PHI Detection", body_font, None),
    ("", "MIT-3.1 CRDT Version Vectors  |  MIT-3.2 Durable Event Sourcing  |  MIT-3.3 Cross-Registry Event Replay", body_font, None),
    ("", "MIT-4.1 Automated Scale Simulation  |  MIT-4.2 Capacity Planning  |  MIT-4.3 Report Tamper-Proofing", body_font, None),
    ("", "MIT-5.1 Zero-Trust Credential Rotation  |  MIT-5.2 Three-Layer Identifier Standard (DID+LEI)  |  MIT-5.3 GDPR Art. 17 Erasure", body_font, None),
    ("", "MIT-6.1 Compliance Dashboard  |  MIT-6.2 Regulatory Change Tracking  |  MIT-6.3 Automated Conformance Reporting", body_font, None),
    ("", "", body_font, None),
    ("", "6 Platform Capabilities:", section_font, None),
    ("", "CAP-1 Multi-Sovereign Federation (IE/GB/US)  |  CAP-2 Policy-Aware Cross-Border Routing", body_font, None),
    ("", "CAP-3 Self-Healing Mesh Network  |  CAP-4 Large-Scale Agent Simulation", body_font, None),
    ("", "CAP-5 Tamper-Evident Audit Trail (Merkle)  |  CAP-6 Persona-Adaptive Clinical Workflows", body_font, None),
    ("", "", body_font, None),
    ("Rationale", "Every mitigation and capability was implemented as a vertical slice across three repositories (GHARRA, Nexus, Integration). A regression failure in any single scenario may indicate: (a) a breaking change in one repo that propagates cross-repo, (b) a Docker topology misconfiguration, (c) a data-seeding gap, or (d) a genuine defect. The 85/10/5 distribution ensures the happy path is thoroughly covered while negative and edge cases validate error handling, security boundaries, and graceful degradation.", section_font, section_fill),
    ("", "", body_font, None),
    ("Test Distribution", "101 Positive (84.2%) -- Expected: all return 2xx with correct response shape", section_font, section_fill),
    ("", "13 Negative (10.8%) -- Expected: all return expected error codes (400/403/404/422/451) with descriptive detail", body_font, None),
    ("", "6 Edge (5.0%) -- Expected: all produce defined, predictable behavior at system boundaries", body_font, None),
    ("", "", body_font, None),
    ("Execution Environment", "Multi-Sovereign Docker Topology (docker-compose.yml):", section_font, section_fill),
    ("", "Root GHARRA (IE):  localhost:8400  --  Primary registry, federation root", body_font, None),
    ("", "GB Sovereign:      localhost:8401  --  NHS/UK GDPR jurisdiction", body_font, None),
    ("", "US Sovereign:      localhost:8402  --  HIPAA jurisdiction", body_font, None),
    ("", "Nexus Gateway:     localhost:8100  --  A2A protocol gateway", body_font, None),
    ("", "SignalBox:         localhost:8221  --  Governance attestation service", body_font, None),
    ("", "", body_font, None),
    ("", "Pre-requisites: All 5 containers healthy. 14 agents seeded across 3 GHARRA instances. Rate limiting disabled (GHARRA_RATE_LIMIT_ENABLED=false).", body_font, None),
    ("", "", body_font, None),
    ("Pass Criteria", "1. All 101 positive scenarios return expected HTTP status and response shape", section_font, section_fill),
    ("", "2. All 13 negative scenarios return the specified error code and error_condition", body_font, None),
    ("", "3. All 6 edge scenarios produce defined, predictable behavior", body_font, None),
    ("", "4. Zero test infrastructure failures (connection refused, timeout, DNS)", body_font, None),
    ("", "5. Ledger hash chain remains valid after all write operations", body_font, None),
    ("", "6. No PHI leakage in any response to non-policy-engine endpoints", body_font, None),
    ("", "7. Cross-registry scenarios confirm federation symmetry (GB and US return equivalent structure to root)", body_font, None),
    ("", "", body_font, None),
    ("Traceability", "Every scenario maps to requirement_ids (FR-*, NFR-*, MIT-*, GDPR-*, CAP-*) enabling full V-model traceability from risk register through implementation to test evidence.", section_font, section_fill),
    ("", "", body_font, None),
    ("Repeatability", "Scenarios are deterministic: given the same Docker topology + seed data, every run produces identical pass/fail results. The JSON matrix (risk_mitigation_regression_matrix.json) is machine-parseable for CI/CD integration.", section_font, section_fill),
]

for ri, (label, value, font, fill) in enumerate(rows, 1):
    c1 = ws3.cell(row=ri, column=1, value=label)
    c2 = ws3.cell(row=ri, column=2, value=value)
    if label:
        c1.font = section_font
        if fill:
            c1.fill = fill
    if ri == 1:
        c2.font = title_font
    else:
        c2.font = font
    if fill and label:
        c2.fill = fill
    c2.alignment = Alignment(wrap_text=True)

wb.save("scenarios/risk_mitigation_regression_matrix.xlsx")
print("DONE: risk_mitigation_regression_matrix.xlsx created with 3 sheets")
